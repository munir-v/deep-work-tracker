# app.py
import rumps
import json
from datetime import datetime
from openpyxl import load_workbook
from Cocoa import (
    NSOpenPanel,
    NSAlert,
    NSComboBox,
    NSPoint,
    NSRect,
    NSSize,
    NSApp,
)
from AppKit import NSAlertFirstButtonReturn
import os
from pathlib import Path

class StopwatchApp(rumps.App):
    def __init__(self):
        super(StopwatchApp, self).__init__("0:00:00")
        self.time_elapsed = 0
        self.timer = rumps.Timer(self.update_time, 1)
        self.running = False

        # Initialize settings
        self.start_at_startup = False
        self.file_location = ""
        self.settings_path = self.get_settings_path()
        self.load_settings()

        # Create menu items
        self.menu = [
            "Start/Resume",
            "Pause",
            "Reset and Save",
            None,
            {
                "Settings": [
                    rumps.MenuItem("Start at startup", callback=self.toggle_startup),
                    rumps.MenuItem("Select .xlsx file location", callback=self.select_file_location),
                ]
            },
        ]

        # Set the initial state of the "Start at startup" setting
        self.menu["Settings"]["Start at startup"].state = self.start_at_startup

    def get_settings_path(self):
        # Get the path to Application Support
        app_support = Path.home() / "Library" / "Application Support" / "StopwatchApp"
        app_support.mkdir(parents=True, exist_ok=True)
        return app_support / "settings.json"

    def update_time(self, _):
        self.time_elapsed += 1
        self.title = self.format_time(self.time_elapsed)

    def format_time(self, seconds):
        hrs = seconds // 3600
        mins = (seconds % 3600) // 60
        secs = seconds % 60
        return f"{hrs}:{mins:02d}:{secs:02d}"

    @rumps.clicked("Start/Resume")
    def start_resume(self, _):
        if not self.running:
            self.timer.start()
            self.running = True

    @rumps.clicked("Pause")
    def pause(self, _):
        if self.running:
            self.timer.stop()
            self.running = False

    @rumps.clicked("Reset and Save")
    def reset_and_save(self, _):
        self.timer.stop()
        self.running = False
        self.save_to_excel()
        self.time_elapsed = 0
        self.title = "0:00:00"

    def toggle_startup(self, sender):
        sender.state = not sender.state
        self.start_at_startup = sender.state
        self.save_settings()
        if self.start_at_startup:
            self.add_to_login_items()
        else:
            self.remove_from_login_items()

    def add_to_login_items(self):
        app_path = os.path.abspath(sys.argv[0])
        script = f'''
        tell application "System Events"
            if not (exists login item "StopwatchApp") then
                make login item at end with properties {{path:"{app_path}", hidden:false}}
            end if
        end tell
        '''
        os.system(f"osascript -e '{script}'")

    def remove_from_login_items(self):
        script = '''
        tell application "System Events"
            delete login item "StopwatchApp"
        end tell
        '''
        os.system(f"osascript -e '{script}'")

    def select_file_location(self, _):
        panel = NSOpenPanel.openPanel()
        panel.setCanChooseFiles_(True)
        panel.setAllowedFileTypes_(["xlsx"])
        panel.setAllowsMultipleSelection_(False)
        if panel.runModal():
            file_url = panel.URLs()[0]
            self.file_location = file_url.path()
            self.save_settings()

    def save_to_excel(self):
        if not self.file_location:
            rumps.alert("No file location selected. Please select a file in Settings.")
            return
        try:
            # Load the Excel file
            file_path = self.file_location
            try:
                workbook = load_workbook(filename=file_path)
            except FileNotFoundError:
                rumps.alert("The selected Excel file does not exist.")
                return

            sheet = workbook.active

            # Get headers from the first row
            headers = [cell.value for cell in sheet[1]]

            # Get a list of columns that do not have headers containing "Date", "Total", or "Other"
            columns = [
                col
                for col in headers
                if col and "Date" not in col and "Total" not in col and "Other" not in col
            ]

            if not columns:
                rumps.alert("No valid categories found in the Excel file.")
                return

            # Present the dropdown list to the user
            category_name = self.select_category(columns)
            if not category_name:
                return  # User canceled the selection

            # Find the column index for the selected category
            column_index = headers.index(category_name) + 1  # openpyxl is 1-indexed

            # Identify the date column (assumed to be the one to the left of the specified column)
            date_column_index = column_index - 1

            # Time value is the elapsed time in minutes
            time_value = round(self.time_elapsed / 60, 2)  # Convert seconds to minutes

            # Find the first empty row in the specified column
            for row in range(2, sheet.max_row + 2):
                if sheet.cell(row=row, column=column_index).value is None:
                    # Insert the current date and time as a datetime object
                    sheet.cell(row=row, column=date_column_index).value = datetime.now()

                    # Insert the time value into the column
                    cell = sheet.cell(row=row, column=column_index)
                    cell.value = time_value
                    cell.number_format = '0.00'  # Set cell format to number with two decimals
                    break

            # Iterate through the sheet to find each "Total" column
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if cell_value and "Total" in cell_value:
                    # Identify the corresponding date column (assumed to be two columns to the left)
                    date_col_idx = col_idx - 2  # Two columns to the left (Date, Value, then Total)
                    if date_col_idx > 0:
                        # Target the first row under the header (row 2) for the sum formula
                        value_range = f"{sheet.cell(row=2, column=col_idx-1).coordinate}:{sheet.cell(row=sheet.max_row, column=col_idx-1).coordinate}"
                        sheet.cell(row=2, column=col_idx).value = f"=SUM({value_range})"

                        # Adding the SUMIFS formula for summing values from the last 7 days
                        date_range = f"{sheet.cell(row=2, column=date_col_idx).coordinate}:{sheet.cell(row=sheet.max_row, column=date_col_idx).coordinate}"
                        sum_previous_week_formula = (
                            f'=SUMIFS({value_range}, {date_range}, ">="&TODAY()-7, {date_range}, "<="&TODAY())'
                        )

                        # Place the sum of the previous week two cells below the total sum (i.e., in row 4)
                        sheet.cell(row=4, column=col_idx).value = sum_previous_week_formula

            # Save the workbook
            workbook.save(filename=file_path)
            rumps.notification("Stopwatch", "Data Saved", f"Time saved under '{category_name}'.")

        except Exception as e:
            rumps.alert(f"Failed to save to Excel file: {e}")

    def select_category(self, categories):
        # Create an NSAlert with an accessory view containing an NSComboBox
        alert = NSAlert.alloc().init()
        alert.setMessageText_("Select Category")
        alert.addButtonWithTitle_("OK")
        alert.addButtonWithTitle_("Cancel")

        # Create the NSComboBox
        view_width = 300
        view_height = 24
        combobox = NSComboBox.alloc().initWithFrame_(
            NSRect(NSPoint(0, 0), NSSize(view_width, view_height))
        )
        combobox.addItemsWithObjectValues_(categories)
        combobox.selectItemAtIndex_(0)  # Select the first item by default

        # Add the combobox to the alert's accessory view
        alert.setAccessoryView_(combobox)

        alert.window().makeKeyAndOrderFront_(None)
        NSApp.activateIgnoringOtherApps_(True)  # Ensures the alert is front-most

        # Set the combobox as the first responder
        alert.window().setInitialFirstResponder_(combobox)

        # Run the alert and get the response
        response = alert.runModal()
        if response == NSAlertFirstButtonReturn:  # OK button
            selected_category = combobox.stringValue()
            return selected_category
        else:
            return None  # User canceled

    def save_settings(self):
        settings = {
            "start_at_startup": self.start_at_startup,
            "file_location": self.file_location,
        }
        with open(self.settings_path, "w") as f:
            json.dump(settings, f)

    def load_settings(self):
        try:
            with open(self.settings_path, "r") as f:
                settings = json.load(f)
                self.start_at_startup = settings.get("start_at_startup", False)
                self.file_location = settings.get("file_location", "")
        except FileNotFoundError:
            pass

if __name__ == "__main__":
    app = StopwatchApp()
    app.run()
